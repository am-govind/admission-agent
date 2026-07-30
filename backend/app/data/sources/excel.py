"""Excel (.xlsx) ingestion source.

Reads whatever tabs the workbook actually has. A tab that is absent is reported as
absent — the previous version of this file invented Finance and Targets rows with
random.random() when they were missing, which meant the agent answered finance
questions with fabricated numbers.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Iterator, Sequence

from ...core.config import settings
from ..schema import TABLE_TAB_SETTING
from .base import SheetSource, SourceError, TableRead

log = logging.getLogger(__name__)

_WINDOW_ROWS = 5000


class ExcelSource(SheetSource):
    name = "excel"

    def __init__(self, path: str | None = None) -> None:
        try:
            import openpyxl
        except ImportError as e:
            raise SourceError("DATA_SOURCE=excel requires openpyxl.") from e

        target = settings.excel_file if path is None else Path(path).expanduser().resolve()
        if not target.exists():
            raise SourceError(f"Excel file not found: {target}")

        self._path = target
        self._wb = openpyxl.load_workbook(target, data_only=True, read_only=True)
        self._titles = set(self._wb.sheetnames)
        log.info("Excel source %s with tabs %s", target, sorted(self._titles))

    def read_table(self, table: str) -> TableRead | None:
        tab = getattr(settings, TABLE_TAB_SETTING[table], "")
        if not tab or tab not in self._titles:
            log.info("Tab %r not present in %s; %s will be unavailable",
                     tab, self._path.name, table)
            return None

        rows = self._wb[tab].iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return None
        return TableRead(header=list(header), windows=self._windows(rows), tab=tab)

    def _windows(self, rows: Iterator[tuple]) -> Iterator[Sequence[Sequence[Any]]]:
        window: list[Sequence[Any]] = []
        for row in rows:
            window.append(list(row))
            if len(window) >= _WINDOW_ROWS:
                yield window
                window = []
        if window:
            yield window

    def close(self) -> None:
        self._wb.close()
