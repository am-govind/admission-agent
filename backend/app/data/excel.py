"""Excel file ingestion path."""
from __future__ import annotations

import csv
import datetime as dt
import os
import tempfile
from pathlib import Path

import openpyxl

from ..core.config import settings
from ..core.database import _lock, get_conn
from .sample_data import _create_rd_table, _gen_finance_rows, _gen_rd_rows, _gen_targets
from .schema import (
    TABLE_FINANCE,
    TABLE_RD25,
    TABLE_RD26,
    TABLE_TARGETS,
)


def load_from_excel(file_path: str | None = None) -> dict[str, int]:
    """Pull sheets from an Excel (.xlsx) file into DuckDB."""
    target_path = file_path or settings.excel_file_path
    if not target_path or not os.path.exists(target_path):
        alt_path = Path(__file__).resolve().parents[2] / "TRY.xlsx"
        if alt_path.exists():
            target_path = str(alt_path)
        else:
            raise FileNotFoundError(f"Excel file not found at: {target_path}")

    wb = openpyxl.load_workbook(target_path, data_only=True, read_only=True)
    sheet_names = wb.sheetnames

    conn = get_conn()
    counts: dict[str, int] = {}
    mapping = {
        TABLE_RD26: settings.tab_rd26,
        TABLE_RD25: settings.tab_rd25,
        TABLE_FINANCE: settings.tab_finance,
        TABLE_TARGETS: settings.tab_targets,
    }

    with _lock:
        for table, tab in mapping.items():
            if tab not in sheet_names:
                if table == TABLE_RD25:
                    rd25_rows = _gen_rd_rows(2025, 4000, seed=25)
                    _create_rd_table(conn, TABLE_RD25, rd25_rows)
                    counts[TABLE_RD25] = len(rd25_rows)
                continue

            ws = wb[tab]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue

            header_list = [str(c).strip().lower() if c is not None else "" for c in header]

            with tempfile.NamedTemporaryFile(mode="w", newline="", encoding="utf-8", delete=False) as tmp:
                writer = csv.writer(tmp)
                writer.writerow(header_list)
                data_count = 0
                for r in rows:
                    if not any(r):
                        continue
                    row_list = list(r)
                    for i, val in enumerate(row_list):
                        if isinstance(val, (dt.date, dt.datetime)):
                            row_list[i] = val.strftime("%d %b, %Y")
                        elif val is None:
                            row_list[i] = ""
                    writer.writerow(row_list)
                    data_count += 1
                tmp_path = tmp.name

            try:
                conn.execute(f"DROP TABLE IF EXISTS {table}")
                conn.execute(
                    f"CREATE TABLE {table} AS SELECT * FROM read_csv_auto('{tmp_path.replace(os.sep, '/')}', all_varchar=True)"
                )
                counts[table] = data_count

                # Standardize numeric columns
                for col_desc in conn.execute(f"SELECT * FROM {table} LIMIT 0").description:
                    col_name = col_desc[0]
                    col_lower = col_name.lower()
                    if (
                        col_lower
                        in {
                            "fees_amt",
                            "fees_paid",
                            "arpu",
                            "pct_discount",
                            "pct_paid",
                            "enrolled_years",
                        }
                        or col_lower.endswith("_target")
                        or col_lower.endswith("_amt")
                        or col_lower.endswith("_paid")
                    ):
                        try:
                            conn.execute(
                                f"UPDATE {table} SET {col_name} = trim(replace(replace(replace(CAST({col_name} AS VARCHAR), ',', ''), '₹', ''), '$', '')) WHERE {col_name} IS NOT NULL"
                            )
                        except Exception:
                            pass
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

        # Generate finance and targets tables if missing
        if TABLE_FINANCE not in counts and TABLE_RD26 in counts:
            rd26_cols = [c[0] for c in conn.execute(f"SELECT * FROM {TABLE_RD26} LIMIT 0").description]
            req_cols = ["region", "batch", "center", "status", "newpayment_checks"]
            select_cols = []
            for col in req_cols:
                if col in rd26_cols:
                    select_cols.append(col)
                else:
                    select_cols.append("''")
            rd26_rows = conn.execute(f"SELECT {', '.join(select_cols)} FROM {TABLE_RD26}").fetchall()
            fin_rows = []
            import random
            for r in rd26_rows:
                scheme = "TRUE" if random.random() < 0.2 else "FALSE"
                fin_rows.append((r[0], r[1], r[2], r[3], r[4], scheme))
            conn.execute(f"DROP TABLE IF EXISTS {TABLE_FINANCE}")
            conn.execute(
                f"CREATE TABLE {TABLE_FINANCE} (region VARCHAR, batch VARCHAR, center VARCHAR, "
                "status VARCHAR, newpayment_checks VARCHAR, vp_ps_regular_schemes VARCHAR)"
            )
            conn.executemany(f"INSERT INTO {TABLE_FINANCE} VALUES (?,?,?,?,?,?)", fin_rows)
            counts[TABLE_FINANCE] = len(fin_rows)

        if TABLE_TARGETS not in counts:
            targets = _gen_targets()
            conn.execute(f"DROP TABLE IF EXISTS {TABLE_TARGETS}")
            conn.execute(
                f"CREATE TABLE {TABLE_TARGETS} (region VARCHAR, center VARCHAR, class_course VARCHAR, "
                "reg_target INTEGER, retention_target INTEGER, monthly_target INTEGER, arpu_target DOUBLE)"
            )
            conn.executemany(f"INSERT INTO {TABLE_TARGETS} VALUES (?,?,?,?,?,?,?)", targets)
            counts[TABLE_TARGETS] = len(targets)

        conn.execute("INSERT OR REPLACE INTO meta VALUES ('last_refresh', ?)", [dt.datetime.now().isoformat()])
        conn.execute("INSERT OR REPLACE INTO meta VALUES ('source', 'excel')")
        conn.execute("INSERT OR REPLACE INTO meta VALUES ('file_path', ?)", [target_path])

    return counts
